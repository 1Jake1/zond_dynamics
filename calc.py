import numpy as np
import pandas as pd
import openpyxl


class DataLoader:
    def __init__(self):
        self.source_data = None
        self.calib_data = None
        self.result_df = None

    def load_excel(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        time_col = []
        tugriki_col = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                time_col.append(row[0])
                tugriki_col.append(row[1])

        n = min(len(time_col), len(tugriki_col))
        self.source_data = pd.DataFrame({
            "Время, мсек": time_col[:n],
            "Тугрики": tugriki_col[:n]
        })

        calib_disp = []
        calib_tugriki = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=5, values_only=True):
            if row[0] is not None and row[1] is not None:
                calib_disp.append(row[0])
                calib_tugriki.append(row[1])

        nc = min(len(calib_disp), len(calib_tugriki))
        self.calib_data = pd.DataFrame({
            "Перемещение, мм": calib_disp[:nc],
            "Тугрики": calib_tugriki[:nc]
        })

        wb.close()
        return self.source_data, self.calib_data

    def _quadratic_interp(self, x, x0, x1, x2, y0, y1, y2):
        """Лагранжевская квадратичная интерполяция по 3 точкам."""
        denom0 = (x0 - x1) * (x0 - x2)
        denom1 = (x1 - x0) * (x1 - x2)
        denom2 = (x2 - x0) * (x2 - x1)

        if denom0 == 0 or denom1 == 0 or denom2 == 0:
            return self._linear_interp(x, x0, x1, y0, y1)

        return (y0 * (x - x1) * (x - x2) / denom0
                + y1 * (x - x0) * (x - x2) / denom1
                + y2 * (x - x0) * (x - x1) / denom2)

    def _linear_interp(self, x, x0, x1, y0, y1):
        """Линейная интерполяция (запасной вариант)."""
        if x1 == x0:
            return y0
        return y0 + (y1 - y0) * (x - x0) / (x1 - x0)

    def calculate(self):
        tugriki_vals = self.source_data["Тугрики"].to_numpy(dtype=float)
        calib_disp = self.calib_data["Перемещение, мм"].to_numpy(dtype=float)
        calib_tugriki = self.calib_data["Тугрики"].to_numpy(dtype=float)

        n_calib = len(calib_tugriki)
        result_disp = np.empty(len(tugriki_vals))

        for i, t in enumerate(tugriki_vals):
            idx = np.argmin(np.abs(calib_tugriki - t))

            if idx == 0:
                i0, i1, i2 = 0, 1, 2
            elif idx == n_calib - 1:
                i0, i1, i2 = n_calib - 3, n_calib - 2, n_calib - 1
            else:
                i0, i1, i2 = idx - 1, idx, idx + 1

            result_disp[i] = self._quadratic_interp(
                t,
                calib_tugriki[i0], calib_tugriki[i1], calib_tugriki[i2],
                calib_disp[i0], calib_disp[i1], calib_disp[i2]
            )

        self.result_df = pd.DataFrame({
            "Время, мсек": self.source_data["Время, мсек"].values,
            "Перемещение, мм": np.round(result_disp, 1)
        })
        return self.result_df
