import sys
import os
import json
import random
import string
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib import rcParams
from pathlib import Path

from calc import DataLoader

BG = "#f0f2f5"
FG = "#1a1a2e"
ACCENT = "#2563eb"
ACCENT_HOVER = "#1d4ed8"
PANEL_BG = "#ffffff"
HEADER_BG = "#e8eaf0"
GRID_COLOR = "#d1d5db"
CHART_BG = "#fafbfc"
SIDEBAR_BG = "#1e293b"
SIDEBAR_FG = "#cbd5e1"
SIDEBAR_ACTIVE = "#334155"


def _style_app():
    style = ttk.Style()
    style.theme_use("clam")

    style.configure(".", background=BG, foreground=FG, font=("Segoe UI", 10))
    style.configure("TFrame", background=BG)
    style.configure("TLabel", background=BG, foreground=FG, font=("Segoe UI", 10))
    style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=(12, 6))
    style.map("TButton",
              background=[("active", ACCENT_HOVER), ("!active", ACCENT)],
              foreground=[("active", "#ffffff"), ("!active", "#ffffff")])

    style.configure("Toolbar.TFrame", background="#1a1a2e")
    style.configure("Toolbar.TButton", background="#2563eb", foreground="#ffffff",
                     font=("Segoe UI", 10, "bold"), padding=(14, 7))
    style.map("Toolbar.TButton",
              background=[("active", "#3b82f6"), ("!active", "#2563eb")],
              foreground=[("active", "#ffffff"), ("!active", "#ffffff")])

    style.configure("Header.TLabel", background="#1a1a2e", foreground="#ffffff",
                     font=("Segoe UI", 11, "bold"))
    style.configure("Status.TLabel", background="#e2e8f0", foreground="#475569",
                     font=("Segoe UI", 9), padding=(10, 4))
    style.configure("Info.TLabel", background=BG, foreground="#64748b",
                     font=("Segoe UI", 9))

    style.configure("TLabelframe", background=PANEL_BG, foreground=FG,
                     font=("Segoe UI", 10, "bold"), relief="solid", borderwidth=1)
    style.configure("TLabelframe.Label", background=PANEL_BG, foreground=ACCENT,
                     font=("Segoe UI", 10, "bold"))

    style.configure("Treeview", background=PANEL_BG, foreground=FG,
                     fieldbackground=PANEL_BG, font=("Consolas", 10), rowheight=24)
    style.configure("Treeview.Heading", background=HEADER_BG, foreground=FG,
                     font=("Segoe UI", 10, "bold"), relief="flat")
    style.map("Treeview", background=[("selected", ACCENT)],
              foreground=[("selected", "#ffffff")])
    style.map("Treeview.Heading", background=[("active", "#d1d5db")])

    style.configure("TPanedwindow", background=BG)
    style.configure("Sash", sashthickness=6, background=GRID_COLOR)

    rcParams["font.family"] = "Segoe UI"
    rcParams["axes.facecolor"] = CHART_BG
    rcParams["figure.facecolor"] = CHART_BG
    rcParams["axes.edgecolor"] = GRID_COLOR
    rcParams["axes.labelcolor"] = FG
    rcParams["xtick.color"] = "#64748b"
    rcParams["ytick.color"] = "#64748b"


class DinamikaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Обработка данных динамики")
        self.root.geometry("1500x850")
        self.root.minsize(1000, 600)
        self.root.state("zoomed")
        self.root.configure(bg=BG)

        self.loader = DataLoader()
        self._file_path = None
        self._history = []
        self._active_idx = None
        self._history_file = Path(os.path.dirname(os.path.abspath(sys.argv[0]))) / "history.json"

        _style_app()
        self._build_ui()
        self._load_history()

    def _build_ui(self):
        self._build_menu()

        outer = ttk.Frame(self.root)
        outer.pack(fill=tk.BOTH, expand=True)

        self._build_sidebar(outer)

        right_frame = ttk.Frame(outer)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._build_toolbar(right_frame)
        self._build_content(right_frame)
        self._build_statusbar()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_menu(self):
        menubar = tk.Menu(self.root, font=("Segoe UI", 10))
        file_menu = tk.Menu(menubar, tearoff=0, font=("Segoe UI", 10))
        file_menu.add_command(label="Открыть Excel...", command=self.load_file, accelerator="Ctrl+O")
        file_menu.add_command(label="Сохранить результат...", command=self.save_file, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.root.quit)
        menubar.add_cascade(label="Файл", menu=file_menu)
        self.root.config(menu=menubar)
        self.root.bind("<Control-o>", lambda e: self.load_file())
        self.root.bind("<Control-s>", lambda e: self.save_file())

    def _build_sidebar(self, parent):
        self.sidebar = tk.Frame(parent, bg=SIDEBAR_BG, width=220)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        header = tk.Label(self.sidebar, text="Файлы", bg=SIDEBAR_BG, fg="#ffffff",
                          font=("Segoe UI", 11, "bold"), anchor="w", padx=10, pady=8)
        header.pack(fill=tk.X)

        sep = tk.Frame(self.sidebar, bg="#475569", height=1)
        sep.pack(fill=tk.X, padx=8)

        list_frame = tk.Frame(self.sidebar, bg=SIDEBAR_BG)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        self.file_listbox = tk.Listbox(list_frame, bg=SIDEBAR_BG, fg=SIDEBAR_FG,
                                       selectbackground=SIDEBAR_ACTIVE,
                                       selectforeground="#ffffff",
                                       font=("Segoe UI", 9),
                                       borderwidth=0, highlightthickness=0,
                                       activestyle="none")
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL,
                                  command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.file_listbox.bind("<<ListboxSelect>>", self._on_file_select)

        btn_frame = tk.Frame(self.sidebar, bg=SIDEBAR_BG)
        btn_frame.pack(fill=tk.X, padx=8, pady=(0, 8))

        self.remove_btn = tk.Button(btn_frame, text="Удалить", bg="#dc2626", fg="#ffffff",
                                    font=("Segoe UI", 9, "bold"), relief="flat",
                                    activebackground="#ef4444", activeforeground="#ffffff",
                                    command=self._remove_from_history)
        self.remove_btn.pack(fill=tk.X, ipady=3)

    def _on_file_select(self, event):
        sel = self.file_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx == self._active_idx:
            return
        self._switch_to_file(idx)

    def _switch_to_file(self, idx):
        entry = self._history[idx]
        self._active_idx = idx
        self._file_path = entry["path"]
        self.loader.source_data = entry["source"]
        self.loader.calib_data = entry["calib"]
        self.loader.result_df = entry["result"]

        self._populate_tree(self.tree_source, self.loader.source_data)
        self._populate_tree(self.tree_calib, self.loader.calib_data)
        self._populate_tree(self.tree_result, self.loader.result_df)

        self._draw_raw_chart()
        self._draw_result_chart()

        self.file_label.configure(text=self._file_path.name)
        src_n = len(self.loader.source_data) if self.loader.source_data is not None else 0
        cal_n = len(self.loader.calib_data) if self.loader.calib_data is not None else 0
        self.status_var.set(f"Загружено: {self._file_path.name}")

        if self.loader.result_df is not None and not self.loader.result_df.empty:
            n = len(self.loader.result_df)
            mn = self.loader.result_df.iloc[:, 1].min()
            mx = self.loader.result_df.iloc[:, 1].max()
            self.stats_var.set(f"Исходных: {src_n}  |  Калибровка: {cal_n}  |  Результат: {n} точек  |  {mn} — {mx} мм")
        else:
            self.stats_var.set(f"Исходных: {src_n}  |  Калибровка: {cal_n} точек")

        self.file_listbox.selection_clear(0, tk.END)
        self.file_listbox.selection_set(idx)

    def _on_close(self):
        self._save_history()
        self.root.destroy()

    def _save_history(self):
        paths = [str(e["path"]) for e in self._history if e["path"].exists()]
        try:
            self._history_file.write_text(json.dumps(paths, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    def _load_history(self):
        if not self._history_file.exists():
            return
        try:
            paths = json.loads(self._history_file.read_text(encoding="utf-8"))
        except Exception:
            return
        for p in paths:
            path = Path(p)
            if path.exists() and path.suffix.lower() == ".xlsx":
                try:
                    loader = DataLoader()
                    loader.load_excel(str(path))
                    loader.calculate()
                    self._history.append({
                        "path": path,
                        "source": loader.source_data.copy() if loader.source_data is not None else None,
                        "calib": loader.calib_data.copy() if loader.calib_data is not None else None,
                        "result": loader.result_df.copy() if loader.result_df is not None else None,
                    })
                except Exception:
                    pass
        self._refresh_file_list()
        if self._history:
            self._switch_to_file(0)

    def _add_to_history(self):
        if self._file_path is None:
            return
        for i, entry in enumerate(self._history):
            if entry["path"] == self._file_path:
                self._history[i] = {
                    "path": self._file_path,
                    "source": self.loader.source_data.copy() if self.loader.source_data is not None else None,
                    "calib": self.loader.calib_data.copy() if self.loader.calib_data is not None else None,
                    "result": self.loader.result_df.copy() if self.loader.result_df is not None else None,
                }
                self._active_idx = i
                self._refresh_file_list()
                self.file_listbox.selection_clear(0, tk.END)
                self.file_listbox.selection_set(i)
                return
        entry = {
            "path": self._file_path,
            "source": self.loader.source_data.copy() if self.loader.source_data is not None else None,
            "calib": self.loader.calib_data.copy() if self.loader.calib_data is not None else None,
            "result": self.loader.result_df.copy() if self.loader.result_df is not None else None,
        }
        self._history.append(entry)
        self._active_idx = len(self._history) - 1
        self._refresh_file_list()
        self.file_listbox.selection_clear(0, tk.END)
        self.file_listbox.selection_set(self._active_idx)

    def _refresh_file_list(self):
        self.file_listbox.delete(0, tk.END)
        for entry in self._history:
            self.file_listbox.insert(tk.END, entry["path"].name)

    def _remove_from_history(self):
        sel = self.file_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        name = self._history[idx]["path"].name
        del self._history[idx]
        self._refresh_file_list()
        if self._active_idx == idx:
            if self._history:
                new_idx = min(idx, len(self._history) - 1)
                self._switch_to_file(new_idx)
            else:
                self._active_idx = None
                self.loader = DataLoader()
                self._file_path = None
                self.tree_source.delete(*self.tree_source.get_children())
                self.tree_calib.delete(*self.tree_calib.get_children())
                self.tree_result.delete(*self.tree_result.get_children())
                self.raw_ax.clear()
                self.raw_canvas.draw()
                self.result_ax.clear()
                self.result_canvas.draw()
                self.file_label.configure(text="Файл не загружен")
                self.status_var.set("Готово")
                self.stats_var.set("")
        elif self._active_idx > idx:
            self._active_idx -= 1

    def _build_toolbar(self, parent):
        tb = ttk.Frame(parent, style="Toolbar.TFrame")
        tb.pack(fill=tk.X, padx=0, pady=0)

        ttk.Label(tb, text="  Динамика", style="Header.TLabel").pack(side=tk.LEFT, padx=(12, 20))

        ttk.Button(tb, text="Открыть", style="Toolbar.TButton",
                   command=self.load_file).pack(side=tk.LEFT, padx=3, pady=6)
        ttk.Button(tb, text="Рассчитать", style="Toolbar.TButton",
                   command=self.calculate).pack(side=tk.LEFT, padx=3, pady=6)
        ttk.Button(tb, text="Сохранить", style="Toolbar.TButton",
                   command=self.save_file).pack(side=tk.LEFT, padx=3, pady=6)

        self.file_label = ttk.Label(tb, text="Файл не загружен", style="Header.TLabel")
        self.file_label.pack(side=tk.RIGHT, padx=15)

    def _build_content(self, parent):
        main_paned = ttk.PanedWindow(parent, orient=tk.VERTICAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(8, 0))

        top_paned = ttk.PanedWindow(main_paned, orient=tk.HORIZONTAL)
        main_paned.add(top_paned, weight=1)

        left = ttk.LabelFrame(top_paned, text=" Динамика — Исходные данные ")
        top_paned.add(left, weight=3)
        self.tree_source = self._make_tree(left)

        right = ttk.LabelFrame(top_paned, text=" Тарировка — Калибровочная кривая ")
        top_paned.add(right, weight=2)
        self.tree_calib = self._make_tree(right)

        raw_chart_frame = ttk.LabelFrame(top_paned, text=" Тугрики от времени ")
        top_paned.add(raw_chart_frame, weight=3)

        self.raw_fig = Figure(figsize=(5, 3), dpi=100)
        self.raw_ax = self.raw_fig.add_subplot(111)
        self.raw_canvas = FigureCanvasTkAgg(self.raw_fig, master=raw_chart_frame)
        self.raw_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        raw_tb = ttk.Frame(raw_chart_frame)
        raw_tb.pack(fill=tk.X, padx=4, pady=(0, 4))
        self.raw_toolbar = NavigationToolbar2Tk(self.raw_canvas, raw_tb)
        self.raw_toolbar.update()

        bot_paned = ttk.PanedWindow(main_paned, orient=tk.HORIZONTAL)
        main_paned.add(bot_paned, weight=3)

        res = ttk.LabelFrame(bot_paned, text=" Результат расчёта ")
        bot_paned.add(res, weight=2)
        self.tree_result = self._make_tree(res)

        chart = ttk.LabelFrame(bot_paned, text=" График: Перемещение от времени ")
        bot_paned.add(chart, weight=3)

        self.result_fig = Figure(figsize=(7, 4), dpi=100)
        self.result_ax = self.result_fig.add_subplot(111)
        self.result_canvas = FigureCanvasTkAgg(self.result_fig, master=chart)
        self.result_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        tb_frame = ttk.Frame(chart)
        tb_frame.pack(fill=tk.X, padx=4, pady=(0, 4))
        self.result_toolbar = NavigationToolbar2Tk(self.result_canvas, tb_frame)
        self.result_toolbar.update()

    def _build_statusbar(self):
        sb = ttk.Frame(self.root, style="Status.TFrame")
        sb.pack(fill=tk.X, side=tk.BOTTOM)

        self.status_var = tk.StringVar(value="Готово")
        ttk.Label(sb, textvariable=self.status_var, style="Status.TLabel").pack(side=tk.LEFT)

        self.stats_var = tk.StringVar(value="")
        ttk.Label(sb, textvariable=self.stats_var, style="Status.TLabel").pack(side=tk.RIGHT)

    def _make_tree(self, parent):
        wrapper = ttk.Frame(parent)
        wrapper.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        search_frame = ttk.Frame(wrapper)
        search_frame.pack(fill=tk.X, pady=(0, 4))

        ttk.Label(search_frame, text="Поиск:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 4))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, font=("Consolas", 9), width=20)
        search_entry.pack(side=tk.LEFT, padx=(0, 4))

        container = ttk.Frame(wrapper)
        container.pack(fill=tk.BOTH, expand=True)

        tree = ttk.Treeview(container, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        tree._df = None
        tree._search_var = search_var
        search_var.trace_add("write", lambda *a, t=tree: self._filter_tree(t))

        return tree

    def _filter_tree(self, tree):
        df = tree._df
        if df is None or df.empty:
            return
        query = tree._search_var.get().strip().lower()
        tree.delete(*tree.get_children())
        cols = list(df.columns)
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=130, minwidth=80, anchor=tk.CENTER)
        for i, (_, row) in enumerate(df.iterrows()):
            vals = [str(v) if pd.notna(v) else "" for v in row]
            if query and not any(query in v.lower() for v in vals):
                continue
            tag = "even" if i % 2 == 0 else "odd"
            tree.insert("", tk.END, values=vals, tags=(tag,))
        tree.tag_configure("even", background="#f8fafc")
        tree.tag_configure("odd", background="#ffffff")

    def _populate_tree(self, tree, df):
        tree._df = df if df is not None else pd.DataFrame()
        tree._search_var.set("")
        self._filter_tree(tree)

    def load_file(self):
        path = filedialog.askopenfilename(
            title="Выберите Excel-файл",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if not path:
            return
        try:
            self.status_var.set("Загрузка файла...")
            self.file_label.configure(text="Загрузка...")
            self.root.update_idletasks()

            self.loader.load_excel(path)
            self._file_path = Path(path)

            self._populate_tree(self.tree_source, self.loader.source_data)
            self._populate_tree(self.tree_calib, self.loader.calib_data)
            self.tree_result.delete(*self.tree_result.get_children())

            self.result_ax.clear()
            self.result_ax.text(0.5, 0.5, "Выполняется расчёт...",
                                ha="center", va="center", transform=self.result_ax.transAxes,
                                fontsize=13, color="#94a3b8", style="italic")
            self.result_ax.set_axis_off()
            self.result_fig.tight_layout()
            self.result_canvas.draw()

            self._draw_raw_chart()

            src_n = len(self.loader.source_data) if self.loader.source_data is not None else 0
            cal_n = len(self.loader.calib_data) if self.loader.calib_data is not None else 0
            self.file_label.configure(text=self._file_path.name)
            self.status_var.set(f"Загружено: {self._file_path.name}")
            self.stats_var.set(f"Исходных: {src_n}  |  Калибровка: {cal_n} точек")

            self.calculate()
            self._add_to_history()
        except Exception as e:
            messagebox.showerror("Ошибка загрузки", str(e))
            self.status_var.set("Ошибка загрузки")
            self.file_label.configure(text="Ошибка")

    def calculate(self):
        if self.loader.source_data is None or self.loader.calib_data is None:
            return

        self.status_var.set("Выполнение расчёта...")
        self.root.update_idletasks()

        try:
            self.loader.calculate()
            self._populate_tree(self.tree_result, self.loader.result_df)
            self._draw_result_chart()

            n = len(self.loader.result_df)
            mn = self.loader.result_df.iloc[:, 1].min()
            mx = self.loader.result_df.iloc[:, 1].max()
            self.status_var.set("Расчёт завершён")

            src_n = len(self.loader.source_data) if self.loader.source_data is not None else 0
            cal_n = len(self.loader.calib_data) if self.loader.calib_data is not None else 0
            self.stats_var.set(f"Исходных: {src_n}  |  Калибровка: {cal_n}  |  Результат: {n} точек  |  {mn} — {mx} мм")
        except Exception as e:
            messagebox.showerror("Ошибка расчёта", str(e))
            self.status_var.set("Ошибка расчёта")

    def _draw_result_chart(self):
        self.result_ax.clear()
        df = self.loader.result_df
        if df is not None and not df.empty:
            self.result_ax.plot(df["Время, мсек"], df["Перемещение, мм"],
                                linewidth=0.8, color="#2196F3")
            self.result_ax.set_xlabel("Время, мсек")
            self.result_ax.set_ylabel("Перемещение, мм")
            self.result_ax.set_title("Перемещение от времени")
            self.result_ax.grid(True, alpha=0.3)
        self.result_fig.tight_layout()
        self.result_canvas.draw()

    def _draw_raw_chart(self):
        self.raw_ax.clear()
        df = self.loader.source_data
        if df is not None and not df.empty:
            self.raw_ax.plot(df["Время, мсек"], df["Тугрики"],
                             linewidth=0.6, color="#4CAF50")
            self.raw_ax.set_xlabel("Время, мсек")
            self.raw_ax.set_ylabel("Тугрики")
            self.raw_ax.set_title("Тугрики от времени")
            self.raw_ax.grid(True, alpha=0.3)
        else:
            self.raw_ax.text(0.5, 0.5, "Нет данных",
                             ha="center", va="center", transform=self.raw_ax.transAxes,
                             fontsize=12, color="#94a3b8")
        self.raw_fig.tight_layout()
        self.raw_canvas.draw()

    def save_file(self):
        if self.loader.result_df is None or self.loader.result_df.empty:
            messagebox.showwarning("Внимание", "Нет данных для сохранения. Сначала выполните расчёт.")
            return

        code = ''.join(random.choices(string.ascii_lowercase + string.digits, k=4))
        if self._file_path:
            stem = self._file_path.stem
        else:
            stem = "Результат"
        default_name = f"{stem}_результат_{code}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Сохранить результат",
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx")],
            initialfile=default_name
        )
        if not path:
            return

        try:
            self.status_var.set("Сохранение...")
            self.root.update_idletasks()

            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = "Результат"

            ws_out["A1"] = "Динамика в тугриках"
            ws_out["D1"] = "Тарировка"
            ws_out["S1"] = "Динамика в мм"

            ws_out["A2"] = "Время, мсек"
            ws_out["B2"] = "Тугрики"
            ws_out["D2"] = "Перемещение, мм"
            ws_out["E2"] = "Тугрики"
            ws_out["S2"] = "Время, мсек"
            ws_out["T2"] = "Перемещение, мм"

            for i, (_, row) in enumerate(self.loader.source_data.iterrows()):
                ws_out.cell(row=i + 3, column=1, value=row["Время, мсек"])
                ws_out.cell(row=i + 3, column=2, value=row["Тугрики"])

            for i, (_, row) in enumerate(self.loader.calib_data.iterrows()):
                ws_out.cell(row=i + 3, column=4, value=row["Перемещение, мм"])
                ws_out.cell(row=i + 3, column=5, value=row["Тугрики"])

            for i, (_, row) in enumerate(self.loader.result_df.iterrows()):
                ws_out.cell(row=i + 3, column=19, value=row["Время, мсек"])
                ws_out.cell(row=i + 3, column=20, value=row["Перемещение, мм"])

            for col, w in {"A": 15, "B": 12, "D": 16, "E": 12, "S": 15, "T": 16}.items():
                ws_out.column_dimensions[col].width = w

            wb_out.save(path)
            wb_out.close()
            self.status_var.set(f"Сохранено: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("Ошибка сохранения", str(e))
            self.status_var.set("Ошибка сохранения")


def main():
    root = tk.Tk()
    app = DinamikaApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
